"""
Excel文件解析器
支持解析不同厂商的PCR数据格式
"""
import pandas as pd
import numpy as np
import openpyxl
from pathlib import Path
import re


class ExcelParser:
    """Excel文件解析器基类"""
    
    def __init__(self):
        self.parsers = {
            'default': DefaultParser(),
            'vendor_a': VendorAParser(),
            'vendor_7500': Vendor7500Parser(),
        }
        
        # 尝试导入VendorBParser（如果存在）
        try:
            from VendorBParser import VendorBParser
            self.parsers['vendor_b'] = VendorBParser()
        except ImportError:
            pass
    
    def parse(self, file_path):
        """解析Excel文件，自动识别格式"""
        # 尝试识别厂商格式
        vendor = self.detect_vendor(file_path)
        
        # 使用对应的解析器
        parser = self.parsers.get(vendor, self.parsers['default'])
        return parser.parse(file_path)
    
    def detect_vendor(self, file_path):
        """检测Excel文件的厂商格式"""
        # 对于.xls文件，使用xlrd读取工作表名
        file_ext = Path(file_path).suffix.lower()
        if file_ext == '.xls':
            try:
                import xlrd
                wb = xlrd.open_workbook(file_path)
                sheet_names = wb.sheet_names()  # 调用方法
            except:
                # 如果xlrd不可用，尝试用openpyxl（可能失败）
                try:
                    wb = openpyxl.load_workbook(file_path)
                    sheet_names = wb.sheetnames
                except:
                    return 'default'
        else:
            wb = openpyxl.load_workbook(file_path)
            sheet_names = wb.sheetnames
        
        # 检测7500格式（必须同时包含多个7500特有的工作表）
        # 7500格式通常包含：Sample Setup, Amplification Data, Results, Raw Data, Multicomponent Data
        has_sample_setup = 'Sample Setup' in sheet_names
        has_amplification = 'Amplification Data' in sheet_names
        has_results = 'Results' in sheet_names
        has_raw_data = 'Raw Data' in sheet_names
        
        # 7500格式至少需要包含Sample Setup和Results，或者包含多个7500特有工作表
        if (has_sample_setup and has_results) or (has_amplification and has_results and has_raw_data):
            return 'vendor_7500'
        
        # 根据工作表名称和内容特征识别
        if any('实验数据' in name or '扩增曲线' in name for name in sheet_names):
            return 'vendor_a'  # 示例：基于中文工作表名
        
        # 可以添加更多识别逻辑
        return 'default'


class BaseParser:
    """解析器基类"""
    
    def parse(self, file_path):
        """解析文件，返回结构化数据"""
        raise NotImplementedError
    
    def extract_experiment_info(self, df):
        """提取实验信息"""
        info = {}
        # 通用提取逻辑
        return info
    
    def extract_amplification_data(self, df):
        """提取扩增数据"""
        raise NotImplementedError


class DefaultParser(BaseParser):
    """默认解析器"""
    
    def parse(self, file_path):
        """解析标准格式的Excel文件"""
        wb = openpyxl.load_workbook(file_path)
        result = {
            'sheets': {},
            'experiment_info': {},
            'amplification_data': pd.DataFrame()
        }
        
        for sheet_name in wb.sheetnames:
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            result['sheets'][sheet_name] = df
            
            # 尝试提取实验信息
            info = self.extract_experiment_info(df)
            result['experiment_info'].update(info)
            
            # 尝试提取扩增数据
            amp_data = self.extract_amplification_data(df)
            if not amp_data.empty:
                result['amplification_data'] = amp_data
        
        return result
    
    def extract_experiment_info(self, df):
        """提取实验信息"""
        info = {}
        
        # 查找关键信息行
        for idx, row in df.iterrows():
            row_str = ' '.join([str(x) for x in row if pd.notna(x)])
            
            # 提取开始时间
            if '开始时间' in row_str or '起始时间' in row_str:
                for i, val in enumerate(row):
                    if pd.notna(val) and i < len(row) - 1:
                        next_val = row[i + 1]
                        if pd.notna(next_val):
                            info['开始时间'] = str(next_val)
                            break
            
            # 提取结束时间
            if '结束时间' in row_str or '完成时间' in row_str:
                for i, val in enumerate(row):
                    if pd.notna(val) and i < len(row) - 1:
                        next_val = row[i + 1]
                        if pd.notna(next_val):
                            info['结束时间'] = str(next_val)
                            break
        
        return info
    
    def extract_amplification_data(self, df):
        """提取扩增数据"""
        # 查找数据区域
        data_start_row = None
        channels = []
        
        # 查找通道名称行（如HEX, CY5, ROX等）
        for idx, row in df.iterrows():
            row_values = [str(x).upper() for x in row if pd.notna(x)]
            
            # 检查是否包含常见通道名
            common_channels = ['HEX', 'CY5', 'ROX', 'FAM', 'VIC', 'CY3']
            found_channels = [ch for ch in common_channels if any(ch in val for val in row_values)]
            
            if found_channels:
                data_start_row = idx
                # 提取通道信息
                for i, val in enumerate(row):
                    if pd.notna(val):
                        val_str = str(val).upper()
                        for ch in common_channels:
                            if ch in val_str:
                                channels.append((i, ch))
                break
        
        if data_start_row is None:
            return pd.DataFrame()
        
        # 提取数据
        data_rows = []
        cycle_col = None
        
        # 查找循环数列
        for idx in range(data_start_row, min(data_start_row + 50, len(df))):
            row = df.iloc[idx]
            first_val = row.iloc[0] if len(row) > 0 else None
            
            # 检查是否是数字（循环数）
            if pd.notna(first_val):
                try:
                    cycle_num = float(first_val)
                    if 1 <= cycle_num <= 50:  # 合理的循环数范围
                        cycle_col = 0
                        break
                except:
                    pass
        
        # 提取数据行
        for idx in range(data_start_row + 1, len(df)):
            row = df.iloc[idx]
            row_data = {}
            
            # 提取循环数
            if cycle_col is not None and pd.notna(row.iloc[cycle_col]):
                try:
                    row_data['Cycle'] = int(float(row.iloc[cycle_col]))
                except:
                    continue
            
            # 提取各通道数据
            for col_idx, channel_name in channels:
                if col_idx < len(row):
                    val = row.iloc[col_idx]
                    if pd.notna(val):
                        val_str = str(val).upper()
                        # 处理NoCt等特殊值
                        if 'NOCT' in val_str or 'N/A' in val_str:
                            row_data[channel_name] = np.nan
                        else:
                            try:
                                row_data[channel_name] = float(val)
                            except:
                                row_data[channel_name] = np.nan
            
            if row_data:
                data_rows.append(row_data)
        
        if data_rows:
            return pd.DataFrame(data_rows)
        return pd.DataFrame()


class VendorAParser(BaseParser):
    """厂商A格式解析器（基于示例文件格式）"""
    
    def parse(self, file_path):
        """解析厂商A的Excel格式"""
        result = {
            'sheets': {},
            'experiment_info': {},
            'amplification_data': pd.DataFrame()
        }
        
        wb = openpyxl.load_workbook(file_path)
        
        # 解析"实验数据"工作表
        if '实验数据' in wb.sheetnames:
            df_exp = pd.read_excel(file_path, sheet_name='实验数据', header=None)
            result['sheets']['实验数据'] = df_exp
            result['experiment_info'] = self.extract_experiment_info(df_exp)
            # 提取孔位数据
            result['well_data'] = self.extract_well_data(df_exp)
            # 从实验数据工作表中提取扩增数据
            result['amplification_data'] = self.extract_amplification_data_from_exp(df_exp)
            # 从实验数据工作表中提取原始数据
            result['raw_data'] = self.extract_raw_data_from_exp(df_exp)
        
        # 解析"扩增曲线"工作表（如果存在）
        if '扩增曲线' in wb.sheetnames:
            df_curve = pd.read_excel(file_path, sheet_name='扩增曲线', header=None)
            result['sheets']['扩增曲线'] = df_curve
            # 如果实验数据中没有扩增数据，则从扩增曲线工作表提取
            if result['amplification_data'].empty:
                result['amplification_data'] = self.extract_amplification_data(df_curve)
        
        return result
    
    def extract_experiment_info(self, df):
        """提取实验信息"""
        info = {}
        
        # 查找关键信息
        for idx in range(min(20, len(df))):
            row = df.iloc[idx]
            first_col = row.iloc[0] if len(row) > 0 else None
            
            if pd.notna(first_col):
                first_str = str(first_col)
                
                # 开始时间
                if '开始时间' in first_str or '起始时间' in first_str:
                    if len(row) > 1 and pd.notna(row.iloc[1]):
                        info['开始时间'] = str(row.iloc[1])
                
                # 结束时间
                if '结束时间' in first_str or '完成时间' in first_str:
                    if len(row) > 1 and pd.notna(row.iloc[1]):
                        info['结束时间'] = str(row.iloc[1])
                
                # 实验名称
                if '实验名称' in first_str:
                    if len(row) > 1 and pd.notna(row.iloc[1]):
                        info['实验名称'] = str(row.iloc[1])
        
        return info
    
    def extract_amplification_data(self, df):
        """从扩增曲线工作表提取数据"""
        # 查找通道行（包含HEX, CY5, ROX等）
        channel_row_idx = None
        channels = []
        
        # 扩大搜索范围
        for idx in range(min(30, len(df))):
            row = df.iloc[idx]
            row_str = ' '.join([str(x).upper() for x in row if pd.notna(x)])
            
            # 检查是否包含通道名
            if any(ch in row_str for ch in ['HEX', 'CY5', 'ROX', 'FAM']):
                channel_row_idx = idx
                # 记录通道位置
                for col_idx, val in enumerate(row):
                    if pd.notna(val):
                        val_str = str(val).upper()
                        if 'HEX' in val_str:
                            channels.append((col_idx, 'HEX'))
                        elif 'CY5' in val_str:
                            channels.append((col_idx, 'CY5'))
                        elif 'ROX' in val_str:
                            channels.append((col_idx, 'ROX'))
                        elif 'FAM' in val_str:
                            channels.append((col_idx, 'FAM'))
                break
        
        # 如果没找到通道行，尝试查找数据区域
        if channel_row_idx is None:
            # 尝试查找包含数字数据的行
            for idx in range(min(30, len(df))):
                row = df.iloc[idx]
                # 检查是否包含数字（可能是数据行）
                numeric_count = 0
                for val in row:
                    if pd.notna(val):
                        try:
                            float(val)
                            numeric_count += 1
                        except:
                            pass
                # 如果一行中有多个数字，可能是数据行
                if numeric_count >= 3:
                    # 假设第一列是循环数，其他列是通道数据
                    channel_row_idx = idx - 1  # 假设上一行是通道名
                    # 尝试从列索引推断通道
                    for col_idx in range(1, min(10, len(row))):
                        # 根据列位置分配通道名（如果找不到通道名）
                        if col_idx == 1:
                            channels.append((col_idx, 'HEX'))
                        elif col_idx == 2:
                            channels.append((col_idx, 'CY5'))
                        elif col_idx == 3:
                            channels.append((col_idx, 'ROX'))
                        elif col_idx == 4:
                            channels.append((col_idx, 'FAM'))
                    break
        
        if channel_row_idx is None or not channels:
            # 如果还是找不到，尝试从数据中推断
            # 查找第一个包含数字的行
            for idx in range(len(df)):
                row = df.iloc[idx]
                first_val = row.iloc[0] if len(row) > 0 else None
                if pd.notna(first_val):
                    try:
                        cycle_num = float(first_val)
                        if 1 <= cycle_num <= 50:  # 合理的循环数
                            channel_row_idx = idx
                            # 假设后续列是通道数据
                            for col_idx in range(1, min(7, len(row))):
                                channels.append((col_idx, ['HEX', 'CY5', 'ROX', 'FAM', 'VIC', 'CY3'][col_idx-1]))
                            break
                    except:
                        pass
        
        if channel_row_idx is None or not channels:
            return pd.DataFrame()
        
        # 提取数据
        data_rows = []
        well_data_map = {}  # 存储每个孔位的数据 {well_name: {cycle: {channel: value}}}
        
        # 从通道行之后开始提取数据
        for idx in range(channel_row_idx + 1, len(df)):
            row = df.iloc[idx]
            row_data = {}
            
            # 尝试提取循环数（通常在某一列）
            cycle_num = idx - channel_row_idx
            if cycle_num > 0:
                row_data['Cycle'] = cycle_num
            
            # 提取各通道的Ct值或荧光值
            for col_idx, channel_name in channels:
                if col_idx < len(row):
                    val = row.iloc[col_idx]
                    if pd.notna(val):
                        val_str = str(val).upper()
                        if 'NOCT' in val_str or 'N/A' in val_str or val_str == 'NAN':
                            row_data[channel_name] = np.nan
                        else:
                            try:
                                row_data[channel_name] = float(val)
                            except:
                                pass
            
            if row_data and 'Cycle' in row_data:
                data_rows.append(row_data)
        
        if data_rows:
            result_df = pd.DataFrame(data_rows)
            # 尝试从实验数据工作表中提取孔位信息
            return result_df
        return pd.DataFrame()
    
    def extract_well_data(self, df_exp):
        """从实验数据工作表中提取孔位数据"""
        well_data = {}
        
        # 查找包含孔位信息的区域
        # 通常孔位信息在表格的某个区域
        for idx in range(len(df_exp)):
            row = df_exp.iloc[idx]
            # 查找包含孔位标识的行（如A1, B2等）
            for col_idx, val in enumerate(row):
                if pd.notna(val):
                    val_str = str(val).strip()
                    # 检查是否是孔位格式（如A1, B12等）
                    if re.match(r'^[A-H][0-9]{1,2}$', val_str, re.IGNORECASE):
                        well_name = val_str.upper()
                        # 尝试提取该孔位的Ct值等信息
                        well_info = {'well': well_name}
                        
                        # 在同一行或相邻行查找Ct值
                        for next_col in range(col_idx + 1, min(col_idx + 10, len(row))):
                            next_val = row.iloc[next_col] if next_col < len(row) else None
                            if pd.notna(next_val):
                                try:
                                    ct_val = float(next_val)
                                    if 0 < ct_val <= 42:  # 合理的Ct值范围（最大42）
                                        well_info['ct'] = ct_val
                                        break
                                except:
                                    pass
                        
                        well_data[well_name] = well_info
        
        return well_data
    
    def extract_amplification_data_from_exp(self, df):
        """从实验数据工作表中提取扩增数据"""
        # 查找表头行（包含"反应孔"、"通道"、"Ct"等）
        header_row_idx = None
        well_col_idx = None
        channel_col_idx = None
        ct_col_idx = None
        sample_name_col_idx = None  # 样本名称列
        data_start_col = None
        
        # 查找表头行（通常是第13行，索引13）
        for idx in range(min(20, len(df))):
            row = df.iloc[idx]
            row_str = ' '.join([str(x) for x in row if pd.notna(x)])
            
            # 查找包含"反应孔"的行（通常是第13行）
            if '反应孔' in row_str:
                header_row_idx = idx
                
                # 直接设置已知的列索引（基于实际数据格式）
                well_col_idx = 0  # 第一列是反应孔
                channel_col_idx = 6  # 第7列是染色（FAM等），第6列是通道编号
                ct_col_idx = 12  # 第13列是Ct
                
                # 查找样本名称列（查找包含"样本"、"样本名称"、"Sample"等关键词的列）
                for col_idx in range(len(row)):
                    if pd.notna(row.iloc[col_idx]):
                        col_str = str(row.iloc[col_idx])
                        if '样本名称' in col_str or '样本' in col_str or 'Sample' in col_str.upper() or '样品名称' in col_str or '样品' in col_str:
                            sample_name_col_idx = col_idx
                            break
                
                # 查找数据开始列（查找包含"1.0"的列，通常是第39列）
                for col_idx in range(35, min(45, len(row))):
                    if pd.notna(row.iloc[col_idx]):
                        val_str = str(row.iloc[col_idx])
                        if val_str == '1.0' or val_str == '1.00':
                            data_start_col = col_idx
                            break
                
                # 如果还是没找到，使用默认值39
                if data_start_col is None:
                    # 检查下一行是否有数据
                    if idx + 1 < len(df):
                        next_row = df.iloc[idx + 1]
                        for col_idx in range(35, min(45, len(next_row))):
                            val = next_row.iloc[col_idx]
                            if pd.notna(val):
                                try:
                                    num_val = float(val)
                                    if -100 < num_val < 10000:
                                        data_start_col = col_idx
                                        break
                                except:
                                    pass
                    if data_start_col is None:
                        data_start_col = 39  # 默认值
                
                break
        
        if header_row_idx is None:
            return pd.DataFrame()
        
        # 提取数据
        data_rows = []
        max_cycles = 0
        
        # 从表头行之后开始提取数据
        for idx in range(header_row_idx + 1, len(df)):
            row = df.iloc[idx]
            
            # 获取孔位
            well_name = None
            if well_col_idx is not None and well_col_idx < len(row):
                well_val = row.iloc[well_col_idx]
                if pd.notna(well_val):
                    well_str = str(well_val).strip()
                    # 检查是否是孔位格式
                    if re.match(r'^[A-H][0-9]{1,2}$', well_str, re.IGNORECASE):
                        well_name = well_str.upper()
            
            if not well_name:
                continue
            
            # 获取样本名称
            sample_name = None
            if sample_name_col_idx is not None and sample_name_col_idx < len(row):
                sample_val = row.iloc[sample_name_col_idx]
                if pd.notna(sample_val):
                    sample_name = str(sample_val).strip()
            
            # 获取通道
            channel_name = None
            if channel_col_idx is not None and channel_col_idx < len(row):
                channel_val = row.iloc[channel_col_idx]
                if pd.notna(channel_val):
                    channel_name = str(channel_val).strip()
            
            # 获取Ct值
            ct_value = None
            if ct_col_idx is not None and ct_col_idx < len(row):
                ct_val = row.iloc[ct_col_idx]
                if pd.notna(ct_val):
                    try:
                        ct_value = float(ct_val)
                    except:
                        pass
            
            # 提取扩增数据（AN列到CC列，即列39到80，共42个循环）
            # AN列索引是40（pandas中索引39），CC列索引是81（pandas中索引80）
            amp_start_col = 39  # AN列
            amp_end_col = 81    # CC列（不包含，所以是81）
            
            if amp_start_col < len(row):
                cycle_num = 1
                for col_idx in range(amp_start_col, min(amp_end_col, len(row))):
                    val = row.iloc[col_idx]
                    if pd.isna(val):
                        # 如果遇到NaN，跳过
                        cycle_num += 1
                        continue
                    
                    try:
                        amp_value = float(val)
                        # 接受合理的扩增值范围（可以是负数，因为可能是ΔRn值）
                        if -100 < amp_value < 10000:
                            row_data = {
                                'Cycle': cycle_num,
                                'Well': well_name,
                                'Channel': channel_name if channel_name else 'Unknown',
                                'Amplification': amp_value
                            }
                            # 只在第一行添加CT值
                            if cycle_num == 1 and ct_value is not None:
                                row_data['Ct'] = ct_value
                            # 每行都添加样本名称（如果存在）
                            if sample_name:
                                row_data['SampleName'] = sample_name
                            data_rows.append(row_data)
                            cycle_num += 1
                            max_cycles = max(max_cycles, cycle_num)
                    except:
                        # 如果转换失败，跳过
                        cycle_num += 1
                        continue
        
        if data_rows:
            result_df = pd.DataFrame(data_rows)
            return result_df
        
        return pd.DataFrame()
    
    def extract_raw_data_from_exp(self, df):
        """从实验数据工作表中提取原始曲线数据"""
        # 查找表头行
        header_row_idx = None
        well_col_idx = None
        channel_col_idx = None
        
        for idx in range(min(20, len(df))):
            row = df.iloc[idx]
            row_str = ' '.join([str(x) for x in row if pd.notna(x)])
            
            if '反应孔' in row_str:
                header_row_idx = idx
                well_col_idx = 0
                channel_col_idx = 6
                break
        
        if header_row_idx is None:
            return pd.DataFrame()
        
        # 提取原始数据（CE列到DT列，即列82到123，共42个循环）
        # CE列索引是83（pandas中索引82），DT列索引是124（pandas中索引123）
        raw_start_col = 82  # CE列
        raw_end_col = 124   # DT列（不包含，所以是124）
        
        data_rows = []
        
        for idx in range(header_row_idx + 1, len(df)):
            row = df.iloc[idx]
            
            # 获取孔位
            well_name = None
            if well_col_idx is not None and well_col_idx < len(row):
                well_val = row.iloc[well_col_idx]
                if pd.notna(well_val):
                    well_str = str(well_val).strip()
                    if re.match(r'^[A-H][0-9]{1,2}$', well_str, re.IGNORECASE):
                        well_name = well_str.upper()
            
            if not well_name:
                continue
            
            # 获取通道
            channel_name = None
            if channel_col_idx is not None and channel_col_idx < len(row):
                channel_val = row.iloc[channel_col_idx]
                if pd.notna(channel_val):
                    channel_name = str(channel_val).strip()
            
            # 提取原始数据
            if raw_start_col < len(row):
                cycle_num = 1
                for col_idx in range(raw_start_col, min(raw_end_col, len(row))):
                    val = row.iloc[col_idx]
                    if pd.isna(val):
                        cycle_num += 1
                        continue
                    
                    try:
                        raw_value = float(val)
                        if -100 < raw_value < 100000:
                            data_rows.append({
                                'Cycle': cycle_num,
                                'Well': well_name,
                                'Channel': channel_name if channel_name else 'Unknown',
                                'RawValue': raw_value
                            })
                            cycle_num += 1
                    except:
                        cycle_num += 1
                        continue
        
        if data_rows:
            result_df = pd.DataFrame(data_rows)
            return result_df
        
        return pd.DataFrame()


class Vendor7500Parser(BaseParser):
    """7500格式解析器（Applied Biosystems 7500）"""
    
    def parse(self, file_path):
        """解析7500格式的Excel文件"""
        result = {
            'sheets': {},
            'experiment_info': {},
            'amplification_data': pd.DataFrame(),
            'raw_data': pd.DataFrame(),
            'well_data': {}
        }
        
        # 根据文件扩展名选择引擎
        file_ext = Path(file_path).suffix.lower()
        if file_ext == '.xls':
            engine = 'xlrd'
        else:
            engine = 'openpyxl'
        
        # 解析Sample Setup工作表
        if self._sheet_exists(file_path, 'Sample Setup', engine):
            df_setup = pd.read_excel(file_path, sheet_name='Sample Setup', header=None, engine=engine)
            result['sheets']['Sample Setup'] = df_setup
            result['experiment_info'] = self.extract_experiment_info(df_setup)
            result['well_data'] = self.extract_well_data_from_setup(df_setup)
        
        # 优先从Multicomponent Data工作表读取扩增数据和原始数据
        if self._sheet_exists(file_path, 'Multicomponent Data', engine):
            df_multicomponent = pd.read_excel(file_path, sheet_name='Multicomponent Data', header=None, engine=engine)
            result['sheets']['Multicomponent Data'] = df_multicomponent
            result['amplification_data'] = self.extract_amplification_data_from_multicomponent(df_multicomponent)
            # 从Multicomponent Data工作表提取原始数据（D列的Rn值）
            result['raw_data'] = self.extract_raw_data_from_multicomponent(df_multicomponent)
        
        # 如果没有Multicomponent Data，则从Amplification Data读取
        if result['amplification_data'].empty and self._sheet_exists(file_path, 'Amplification Data', engine):
            df_amp = pd.read_excel(file_path, sheet_name='Amplification Data', header=None, engine=engine)
            result['sheets']['Amplification Data'] = df_amp
            result['amplification_data'] = self.extract_amplification_data(df_amp)
        
        # 解析Results工作表（获取Ct值）
        if self._sheet_exists(file_path, 'Results', engine):
            df_results = pd.read_excel(file_path, sheet_name='Results', header=None, engine=engine)
            result['sheets']['Results'] = df_results
            # 从Results中提取Ct值并更新到well_data
            ct_data = self.extract_ct_from_results(df_results)
            for well_name, channel_ct in ct_data.items():
                if well_name not in result['well_data']:
                    result['well_data'][well_name] = {}
                result['well_data'][well_name].update(channel_ct)
        
        # 如果没有从Multicomponent Data获取原始数据，则从Raw Data工作表读取
        if result['raw_data'].empty and self._sheet_exists(file_path, 'Raw Data', engine):
            df_raw = pd.read_excel(file_path, sheet_name='Raw Data', header=None, engine=engine)
            result['sheets']['Raw Data'] = df_raw
            result['raw_data'] = self.extract_raw_data(df_raw)
        
        return result
    
    def _sheet_exists(self, file_path, sheet_name, engine):
        """检查工作表是否存在"""
        try:
            if engine == 'xlrd':
                import xlrd
                wb = xlrd.open_workbook(file_path)
                return sheet_name in wb.sheet_names()  # 调用方法
            else:
                wb = openpyxl.load_workbook(file_path)
                return sheet_name in wb.sheetnames
        except:
            return False
    
    def extract_experiment_info(self, df):
        """提取实验信息"""
        info = {}
        
        # 查找关键信息（前7行）
        for idx in range(min(7, len(df))):
            row = df.iloc[idx]
            if len(row) > 0 and pd.notna(row.iloc[0]):
                key = str(row.iloc[0]).strip()
                if len(row) > 1 and pd.notna(row.iloc[1]):
                    value = str(row.iloc[1]).strip()
                    if key and value:
                        info[key] = value
        
        return info
    
    def extract_well_data_from_setup(self, df):
        """从Sample Setup工作表提取孔位和通道信息"""
        well_data = {}
        
        # 查找表头行（通常是第7行，索引7）
        header_row = None
        for idx in range(min(10, len(df))):
            row = df.iloc[idx]
            row_str = ' '.join([str(x) for x in row if pd.notna(x)])
            if 'Well' in row_str and 'Target Name' in row_str:
                header_row = idx
                break
        
        if header_row is None:
            return well_data
        
        # 确定列索引
        header = df.iloc[header_row]
        well_col = None
        target_col = None
        sample_name_col = None
        
        for i, val in enumerate(header):
            if pd.notna(val):
                val_str = str(val).strip()
                if val_str == 'Well':
                    well_col = i
                elif val_str == 'Target Name':
                    target_col = i
                elif val_str == 'Sample Name':
                    sample_name_col = i
        
        if well_col is None or target_col is None:
            return well_data
        
        # 提取数据（从表头行之后开始）
        for idx in range(header_row + 1, len(df)):
            row = df.iloc[idx]
            
            # 获取孔位
            if well_col < len(row) and pd.notna(row.iloc[well_col]):
                well_name = str(row.iloc[well_col]).strip()
                if re.match(r'^[A-H][0-9]{1,2}$', well_name, re.IGNORECASE):
                    well_name = well_name.upper()
                    
                    # 获取通道名（Target Name）
                    if target_col < len(row) and pd.notna(row.iloc[target_col]):
                        target_name = str(row.iloc[target_col]).strip()
                        
                        # 映射通道名（HEX -> VIC, JOE -> VIC）
                        if target_name == 'HEX' or target_name == 'JOE':
                            channel_name = 'VIC'
                        else:
                            channel_name = target_name
                        
                        # 获取样本名称
                        sample_name = None
                        if sample_name_col is not None and sample_name_col < len(row):
                            sample_val = row.iloc[sample_name_col]
                            if pd.notna(sample_val):
                                sample_name = str(sample_val).strip()
                        
                        if well_name not in well_data:
                            well_data[well_name] = {}
                        
                        if 'channels' not in well_data[well_name]:
                            well_data[well_name]['channels'] = []
                        
                        if channel_name not in well_data[well_name]['channels']:
                            well_data[well_name]['channels'].append(channel_name)
                        
                        if sample_name and 'sample_name' not in well_data[well_name]:
                            well_data[well_name]['sample_name'] = sample_name
        
        return well_data
    
    def extract_amplification_data(self, df):
        """从Amplification Data工作表提取扩增数据"""
        # 查找表头行（通常是第7行，索引7）
        header_row = None
        for idx in range(min(10, len(df))):
            row = df.iloc[idx]
            row_str = ' '.join([str(x) for x in row if pd.notna(x)])
            if 'Well' in row_str and 'Cycle' in row_str:
                header_row = idx
                break
        
        if header_row is None:
            return pd.DataFrame()
        
        # 确定列索引
        header = df.iloc[header_row]
        well_col = None
        cycle_col = None
        target_col = None
        rn_col = None
        delta_rn_col = None
        
        for i, val in enumerate(header):
            if pd.notna(val):
                val_str = str(val).strip()
                if val_str == 'Well':
                    well_col = i
                elif val_str == 'Cycle':
                    cycle_col = i
                elif val_str == 'Target Name':
                    target_col = i
                elif val_str == 'Rn' or val_str == 'Rn':
                    rn_col = i
                elif 'ΔRn' in val_str or 'Delta Rn' in val_str or 'dRn' in val_str:
                    delta_rn_col = i
        
        if well_col is None or cycle_col is None:
            return pd.DataFrame()
        
        # 提取数据
        data_rows = []
        for idx in range(header_row + 1, len(df)):
            row = df.iloc[idx]
            
            # 获取孔位
            if well_col >= len(row) or pd.isna(row.iloc[well_col]):
                continue
            
            well_name = str(row.iloc[well_col]).strip()
            if not re.match(r'^[A-H][0-9]{1,2}$', well_name, re.IGNORECASE):
                continue
            well_name = well_name.upper()
            
            # 获取循环数
            if cycle_col >= len(row) or pd.isna(row.iloc[cycle_col]):
                continue
            
            try:
                cycle = int(float(row.iloc[cycle_col]))
            except:
                continue
            
            # 获取通道名
            channel_name = None
            if target_col is not None and target_col < len(row) and pd.notna(row.iloc[target_col]):
                target_name = str(row.iloc[target_col]).strip()
                # 保留HEX作为独立通道，不映射为VIC（因为UI中有HEX选项）
                # JOE映射为VIC（JOE是VIC的旧名称）
                if target_name == 'JOE':
                    channel_name = 'VIC'
                else:
                    channel_name = target_name  # HEX保持为HEX
            
            if not channel_name:
                continue
            
            # 获取扩增值（优先使用ΔRn，如果没有则使用Rn）
            amp_value = None
            if delta_rn_col is not None and delta_rn_col < len(row) and pd.notna(row.iloc[delta_rn_col]):
                try:
                    amp_value = float(row.iloc[delta_rn_col])
                except:
                    pass
            
            if amp_value is None and rn_col is not None and rn_col < len(row) and pd.notna(row.iloc[rn_col]):
                try:
                    amp_value = float(row.iloc[rn_col])
                except:
                    pass
            
            if amp_value is not None:
                data_rows.append({
                    'Cycle': cycle,
                    'Well': well_name,
                    'Channel': channel_name,
                    'Amplification': amp_value
                })
        
        if data_rows:
            return pd.DataFrame(data_rows)
        return pd.DataFrame()
    
    def extract_ct_from_results(self, df):
        """从Results工作表提取Ct值"""
        ct_data = {}
        
        # 查找表头行（通常是第7行，索引7）
        header_row = None
        for idx in range(min(10, len(df))):
            row = df.iloc[idx]
            row_str = ' '.join([str(x) for x in row if pd.notna(x)])
            if 'Well' in row_str and 'Target Name' in row_str:
                header_row = idx
                break
        
        if header_row is None:
            return ct_data
        
        # 确定列索引
        header = df.iloc[header_row]
        well_col = None
        target_col = None
        ct_col = None
        
        for i, val in enumerate(header):
            if pd.notna(val):
                val_str = str(val).strip()
                if val_str == 'Well':
                    well_col = i
                elif val_str == 'Target Name':
                    target_col = i
                # 对于7500格式，Ct值列固定在第6列（G列，索引6），不通过列名匹配
                # 因为列名可能有编码问题
        
        # 对于7500格式，Ct值列固定在第6列（G列，索引6）
        if len(header) > 6:
            ct_col = 6  # G列，索引6
        else:
            return ct_data
        
        if well_col is None or target_col is None:
            return ct_data
        
        # 提取数据
        ct_count = 0
        for idx in range(header_row + 1, len(df)):
            row = df.iloc[idx]
            
            # 获取孔位
            if well_col >= len(row) or pd.isna(row.iloc[well_col]):
                continue
            
            well_name = str(row.iloc[well_col]).strip()
            if not re.match(r'^[A-H][0-9]{1,2}$', well_name, re.IGNORECASE):
                continue
            well_name = well_name.upper()
            
            # 获取通道名
            if target_col >= len(row) or pd.isna(row.iloc[target_col]):
                continue
            
            target_name = str(row.iloc[target_col]).strip()
            # 映射通道名
            if target_name == 'HEX' or target_name == 'JOE':
                channel_name = 'VIC'
            else:
                channel_name = target_name
            
            # 获取Ct值（从第6列，G列）
            if ct_col < len(row) and pd.notna(row.iloc[ct_col]):
                ct_val = row.iloc[ct_col]
                # 处理"Undetermined"、"N"、"N/A"等特殊值
                if isinstance(ct_val, str):
                    ct_val_str = ct_val.strip().upper()
                    if ct_val_str in ['UNDETERMINED', 'N/A', 'N', 'NA', '']:
                        continue
                
                try:
                    ct_value = float(ct_val)
                    if 0 < ct_value <= 42:  # 合理的Ct值范围（最大42）
                        if well_name not in ct_data:
                            ct_data[well_name] = {}
                        ct_data[well_name][channel_name] = ct_value
                        ct_count += 1
                except Exception as e:
                    pass
        
        return ct_data
    
    def extract_amplification_data_from_multicomponent(self, df):
        """从Multicomponent Data工作表提取扩增数据"""
        # 查找表头行（通常是第7行，索引7）
        header_row = None
        for idx in range(min(10, len(df))):
            row = df.iloc[idx]
            row_str = ' '.join([str(x) for x in row if pd.notna(x)])
            if 'Well' in row_str and 'Cycle' in row_str:
                header_row = idx
                break
        
        if header_row is None:
            return pd.DataFrame()
        
        # 确定列索引
        header = df.iloc[header_row]
        well_col = None
        cycle_col = None
        target_col = None  # Target Name列，用于确定通道
        delta_rn_col = 4  # E列，索引4（固定使用E列作为delta Rn值）
        
        for i, val in enumerate(header):
            if pd.notna(val):
                val_str = str(val).strip()
                if val_str == 'Well':
                    well_col = i
                elif val_str == 'Cycle':
                    cycle_col = i
                elif val_str == 'Target Name':
                    target_col = i
        
        if well_col is None or cycle_col is None or target_col is None:
            return pd.DataFrame()
        
        # 提取数据，直接使用E列的Delta Rn值
        data_rows = []
        
        for idx in range(header_row + 1, len(df)):
            row = df.iloc[idx]
            
            # 获取孔位
            if well_col >= len(row) or pd.isna(row.iloc[well_col]):
                continue
            
            well_name = str(row.iloc[well_col]).strip()
            if not re.match(r'^[A-H][0-9]{1,2}$', well_name, re.IGNORECASE):
                continue
            well_name = well_name.upper()
            
            # 获取循环数
            if cycle_col >= len(row) or pd.isna(row.iloc[cycle_col]):
                continue
            
            try:
                cycle = int(float(row.iloc[cycle_col]))
            except:
                continue
            
            # 获取通道名（从Target Name列）
            if target_col >= len(row) or pd.isna(row.iloc[target_col]):
                continue
            
            target_name = str(row.iloc[target_col]).strip()
            # 映射通道名：JOE -> VIC, HEX保持为HEX
            if target_name == 'JOE':
                channel_name = 'VIC'
            elif target_name == 'HEX':
                channel_name = 'HEX'  # 保留HEX作为独立通道
            else:
                channel_name = target_name
            
            # 获取E列的delta Rn值（索引4）
            if delta_rn_col >= len(row) or pd.isna(row.iloc[delta_rn_col]):
                continue
            
            try:
                delta_rn_value = float(row.iloc[delta_rn_col])
            except:
                continue
            
            data_rows.append({
                'Cycle': cycle,
                'Well': well_name,
                'Channel': channel_name,
                'Amplification': delta_rn_value
            })
        
        if data_rows:
            result_df = pd.DataFrame(data_rows)
            
            # Debug: 输出C2孔位FAM通道的最终结果汇总
            c2_fam_data = result_df[(result_df['Well'] == 'C2') & (result_df['Channel'] == 'FAM')]
            return result_df
        return pd.DataFrame()
    
    def extract_raw_data_from_multicomponent(self, df):
        """从Multicomponent Data工作表提取原始数据（D列的Rn值）"""
        # 查找表头行（通常是第7行，索引7）
        header_row = None
        for idx in range(min(10, len(df))):
            row = df.iloc[idx]
            row_str = ' '.join([str(x) for x in row if pd.notna(x)])
            if 'Well' in row_str and 'Cycle' in row_str:
                header_row = idx
                break
        
        if header_row is None:
            return pd.DataFrame()
        
        # 确定列索引
        header = df.iloc[header_row]
        well_col = None
        cycle_col = None
        channel_cols = {}  # {channel_name: col_index} 通道列映射
        rn_col = 3  # D列，索引3（固定使用D列作为Rn值）
        
        for i, val in enumerate(header):
            if pd.notna(val):
                val_str = str(val).strip()
                if val_str == 'Well':
                    well_col = i
                elif val_str == 'Cycle':
                    cycle_col = i
                elif val_str in ['FAM', 'JOE', 'CY5', 'ROX', 'VIC', 'HEX']:
                    # 映射通道名：JOE -> VIC, HEX保持为HEX
                    if val_str == 'JOE':
                        channel_name = 'VIC'
                    elif val_str == 'HEX':
                        channel_name = 'HEX'  # 保留HEX作为独立通道
                    else:
                        channel_name = val_str
                    channel_cols[channel_name] = i
        
        if well_col is None or cycle_col is None or not channel_cols:
            return pd.DataFrame()
        
        # 提取数据，直接使用D列的Rn值
        data_rows = []
        processed_count = 0
        skipped_count = 0
        
        for idx in range(header_row + 1, len(df)):
            row = df.iloc[idx]
            processed_count += 1
            
            # 获取孔位
            if well_col >= len(row) or pd.isna(row.iloc[well_col]):
                skipped_count += 1
                continue
            
            well_name = str(row.iloc[well_col]).strip()
            if not re.match(r'^[A-H][0-9]{1,2}$', well_name, re.IGNORECASE):
                skipped_count += 1
                continue
            well_name = well_name.upper()
            
            # 获取循环数
            if cycle_col >= len(row) or pd.isna(row.iloc[cycle_col]):
                skipped_count += 1
                continue
            
            try:
                cycle = int(float(row.iloc[cycle_col]))
            except:
                skipped_count += 1
                continue
            
            # 获取D列的Rn值（索引3）
            if rn_col >= len(row) or pd.isna(row.iloc[rn_col]):
                skipped_count += 1
                continue
            
            try:
                rn_value = float(row.iloc[rn_col])
            except:
                skipped_count += 1
                continue
            
            # 确定通道名：检查哪个通道列有值（除了Well、Cycle和D列）
            # 如果某个通道列有值，说明这一行属于该通道
            channel_name = None
            for ch_name, ch_col_idx in channel_cols.items():
                if ch_col_idx < len(row) and pd.notna(row.iloc[ch_col_idx]):
                    # 检查该通道列是否有有效值
                    try:
                        ch_value = float(row.iloc[ch_col_idx])
                        # 如果通道列有值，说明这一行属于该通道
                        channel_name = ch_name
                        break
                    except:
                        pass
            
            # 如果无法确定通道，跳过
            if not channel_name:
                skipped_count += 1
                continue
            
            data_rows.append({
                'Cycle': cycle,
                'Well': well_name,
                'Channel': channel_name,
                'RawValue': rn_value
            })
        
        if data_rows:
            result_df = pd.DataFrame(data_rows)
            return result_df
        return pd.DataFrame()
    
    def extract_raw_data(self, df):
        """从Raw Data工作表提取原始数据"""
        # 查找表头行（通常是第7行，索引7）
        header_row = None
        for idx in range(min(10, len(df))):
            row = df.iloc[idx]
            row_str = ' '.join([str(x) for x in row if pd.notna(x)])
            if 'Well' in row_str and 'Cycle' in row_str:
                header_row = idx
                break
        
        if header_row is None:
            return pd.DataFrame()
        
        # 确定列索引
        header = df.iloc[header_row]
        well_col = None
        cycle_col = None
        
        for i, val in enumerate(header):
            if pd.notna(val):
                val_str = str(val).strip()
                if val_str == 'Well':
                    well_col = i
                elif val_str == 'Cycle':
                    cycle_col = i
        
        if well_col is None or cycle_col is None:
            return pd.DataFrame()
        
        # 确定通道列（从cycle_col之后开始，每列是一个通道的原始值）
        # 需要从Sample Setup获取通道信息，这里简化处理，假设列顺序对应通道
        # 实际应该从Sample Setup获取每个孔位的通道配置
        
        # 提取数据
        data_rows = []
        for idx in range(header_row + 1, len(df)):
            row = df.iloc[idx]
            
            # 获取孔位
            if well_col >= len(row) or pd.isna(row.iloc[well_col]):
                continue
            
            well_name = str(row.iloc[well_col]).strip()
            if not re.match(r'^[A-H][0-9]{1,2}$', well_name, re.IGNORECASE):
                continue
            well_name = well_name.upper()
            
            # 获取循环数
            if cycle_col >= len(row) or pd.isna(row.iloc[cycle_col]):
                continue
            
            try:
                cycle = int(float(row.iloc[cycle_col]))
            except:
                continue
            
            # 从cycle_col之后开始，每列是一个通道的原始值
            # 这里需要知道通道顺序，暂时跳过，因为Raw Data的格式比较复杂
            # 可以根据需要后续完善
        
        if data_rows:
            return pd.DataFrame(data_rows)
        return pd.DataFrame()

